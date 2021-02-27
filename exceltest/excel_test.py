# -*- coding: utf-8 -*-
'''
@Time    : 2021/2/27 11:35
@Author  : zxd
'''
'''
功能描述:
  读取:
    1.打开excel
    2.确定sheet页
    3.确定数据行和列
    4.读取一行数据
    5.读取所有行数据
    
    
   写入:xlutils.copy
'''
import xlrd
from xlutils.copy import copy
import openpyxl
#读取
#1.打开excel
ec = xlrd.open_workbook('D:\\apitest\\VIP9Interface_framework\\testData\\data.xls')

#2.确定sheet页
st = ec.sheet_by_index(0)
#3.获取总行数和总列数
nr = st.nrows   #行数
nc = st.ncols  #列数
print(nr,nc)
for i in range(0,nr):
    print(st.row_values(i))
# #4.输出logout值  cell_value(x,y) 第x+1行,y+1列
# logout_print = st.cell_value(3,2)
# print('logout_print的值为',logout_print)
# #5.获取所有行和列的值
# row_values = st.row_values(1)
# print(row_values)

# #写入
# #1.复制原来的excel变为一个新的excel
# n_exc = copy(ec)
# #2.获取sheet
# st=n_exc.get_sheet(1)
# #3.写入
# st.write(1,1,'hello world')
# #4.保存
# n_exc.save('D:\\apitest\\VIP9Interface_framework\\testData\\data.xls')

# #openpyxl方法
#
# filename = 'D:\\apitest\\VIP9Interface_framework\\testData\\test.xlsx'
# wb2 = openpyxl.load_workbook(filename)
# st_names = wb2.get_sheet_names()
# st=wb2.get_sheet_by_name(st_names[0])
#
# rows = st.max_row
# clos = st.max_column
# print(rows,clos)
# print(st.cell(2,2).value)  #cell(x,y) 第x行,第y列
